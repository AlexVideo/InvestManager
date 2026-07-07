# export_excel.py
# Экспорт данных в Excel (.xlsx): сводная + лист на каждый проект.
# Требуется: pip install openpyxl

from __future__ import annotations
import os
import re
import datetime as _dt

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

import db

_MAX_SHEETNAME = 31
_INVALID_SHEET_CHARS = re.compile(r'[\\*?:/\[\]]')
_NUM_FMT_MONEY = '# ##0,00'
_NUM_FMT_PERCENT = '0.0%'

def _sanitize_sheet_name(name: str) -> str:
    """Excel запрещает в имени листа: \\ / ? * [ ] :"""
    cleaned = _INVALID_SHEET_CHARS.sub("-", (name or "").strip())
    cleaned = cleaned[:_MAX_SHEETNAME].strip(" .")
    return cleaned or "Лист"

def _uniq_sheet_name(base: str, used: set[str]) -> str:
    base = _sanitize_sheet_name(base)
    name = base[:_MAX_SHEETNAME] if len(base) > _MAX_SHEETNAME else base
    if name not in used:
        used.add(name); return name
    i = 2
    while True:
        suffix = f"_{i}"
        cut = _MAX_SHEETNAME - len(suffix)
        cand = (base[:cut] if len(base) > cut else base) + suffix
        if cand not in used:
            used.add(cand); return cand
        i += 1

def _autosize(ws):
    widths = {}
    for row in ws.rows:
        for cell in row:
            v = str(cell.value) if cell.value is not None else ""
            widths[cell.column] = max(widths.get(cell.column, 0), len(v))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, w + 2), 60)

def _style_header_row(ws, row: int, ncol: int, head_fill, head_font):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = head_fill
        cell.font = head_font

def _write_summary_cell(ws, row: int, col: int, value, *, money: bool = False, percent: bool = False, right=None):
    cell = ws.cell(row=row, column=col, value=value)
    if money and isinstance(value, (int, float)):
        cell.number_format = _NUM_FMT_MONEY
        if right:
            cell.alignment = right
    elif percent and isinstance(value, (int, float)):
        cell.number_format = _NUM_FMT_PERCENT
        if right:
            cell.alignment = right
    elif right and isinstance(value, (int, float)):
        cell.alignment = right
    return cell

def _create_project_sheet(
    wb,
    pid: int,
    name: str,
    used_names: set[str],
    sheet_map: dict[int, str],
    head_fill,
    head_font,
    right,
):
    sheet_name = _uniq_sheet_name(name if name else f"Проект_{pid}", used_names)
    sheet_map[pid] = sheet_name
    ws = wb.create_sheet(title=sheet_name)

    ws["A1"] = f"Карточка проекта: {name}"
    ws["A1"].font = Font(bold=True, size=14)

    pr = db.get_project(pid)
    base = float(pr[2]) if pr else 0.0
    out_of_budget = bool(pr[5]) if pr and len(pr) > 5 else False
    st = db.compute_project_status(pid)

    summary = [
        ("Выделено", base, True),
        ("Имеется", st["have"], True),
        ("Необходимо", st["need"], True),
        ("Остаток", st["diff"], True),
        ("Вне бюджета", "Да" if out_of_budget else "Нет", False),
        ("Рудник", db.get_mine_name(pr[6]) if pr and len(pr) > 6 and pr[6] else "—", False),
        ("Участок", db.get_section_name(pr[7]) if pr and len(pr) > 7 and pr[7] else "—", False),
    ]
    for i, (label, val, is_money) in enumerate(summary, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=val)
        if is_money:
            cell.number_format = _NUM_FMT_MONEY
            cell.alignment = right

    ws["A11"] = "История по датам"
    ws["A11"].font = Font(bold=True)
    hist_headers = ["Дата", "Тип", "Сумма", "Комментарий", "Файл"]
    for c, h in enumerate(hist_headers, start=1):
        ws.cell(row=12, column=c, value=h)
    _style_header_row(ws, 12, len(hist_headers), head_fill, head_font)

    r = 13
    for ev in db.get_project_timeline(pid):
        ws.cell(row=r, column=1, value=ev["date"])
        ws.cell(row=r, column=2, value=ev["type"])
        amount = ev.get("amount")
        if amount is not None:
            amt_cell = ws.cell(row=r, column=3, value=float(amount))
            amt_cell.number_format = _NUM_FMT_MONEY
            amt_cell.alignment = right
        else:
            ws.cell(row=r, column=3, value="")
        ws.cell(row=r, column=4, value=ev.get("note") or "")
        ws.cell(row=r, column=5, value=ev.get("file_path") or "")
        r += 1

    _autosize(ws)
    return sheet_name

def export_table_to_excel(
    xlsx_path: str,
    headers: list[str],
    rows: list[list],
    project_ids: list[int] | None = None,
    *,
    numeric_col_indexes: list[int] | None = None,
    percent_col_indexes: list[int] | None = None,
    name_col_index: int = 0,
) -> str:
    """
    Экспортирует таблицу в Excel:
      - Лист «Сводная» с числами как числами и гиперссылками на проекты
      - По листу на каждый проект из project_ids (хронология действий)
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("Не установлен пакет 'openpyxl'. Установите: pip install openpyxl")

    numeric_col_indexes = numeric_col_indexes or []
    percent_col_indexes = percent_col_indexes or []
    project_ids = project_ids or []

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводная"

    head_fill = PatternFill("solid", fgColor="333333")
    head_font = Font(bold=True, color="FFFFFF")
    right = Alignment(horizontal="right")

    ws_sum.append(headers)
    _style_header_row(ws_sum, 1, len(headers), head_fill, head_font)

    used_names: set[str] = {"Сводная"}
    sheet_map: dict[int, str] = {}

    for row_idx, row in enumerate(rows, start=2):
        pid = project_ids[row_idx - 2] if row_idx - 2 < len(project_ids) else None
        if pid and pid not in sheet_map:
            name = ""
            if 0 <= name_col_index < len(row) and row[name_col_index] is not None:
                name = str(row[name_col_index])
            _create_project_sheet(wb, pid, name, used_names, sheet_map, head_fill, head_font, right)

        for col_idx, value in enumerate(row, start=1):
            is_money = (col_idx - 1) in numeric_col_indexes
            is_percent = (col_idx - 1) in percent_col_indexes
            _write_summary_cell(
                ws_sum, row_idx, col_idx, value,
                money=is_money, percent=is_percent, right=right,
            )

        if pid and pid in sheet_map and 0 <= name_col_index < len(row):
            link_cell = ws_sum.cell(row=row_idx, column=name_col_index + 1)
            target_sheet = sheet_map[pid]
            link_cell.hyperlink = f"#'{target_sheet}'!A1"
            link_cell.style = "Hyperlink"

    _autosize(ws_sum)
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)) or ".", exist_ok=True)
    wb.save(xlsx_path)
    return os.path.abspath(xlsx_path)


def export_to_excel(xlsx_path: str) -> str:
    """
    Экспортирует все проекты в Excel:
      - Лист 'Сводная' (Название, Выделено, Имеется, Необходимо, Остаток, Статус) + гиперссылки на листы проектов
      - По листу на каждый проект (сводка + история)
    Возвращает абсолютный путь к файлу. Бросает RuntimeError, если нет openpyxl.
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("Не установлен пакет 'openpyxl'. Установите: pip install openpyxl")

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Сводная"

    # Стили
    head_fill = PatternFill("solid", fgColor="333333")
    head_font = Font(bold=True, color="FFFFFF")
    right = Alignment(horizontal="right")

    # Заголовок сводной
    headers = ["Название", "Выделено", "Имеется", "Необходимо", "Остаток", "Статус", "Вне бюджета", "Рудник", "Участок"]
    ws_sum.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws_sum.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font

    # Листы проектов
    used_names: set[str] = {"Сводная"}
    summary_rows = []

    projects = db.list_projects()
    for pid, name, base_budget, comment, created_at, out_of_budget, mine_id, section_id, _ in projects:
        st = db.compute_project_status(pid)
        mine_name = db.get_mine_name(mine_id) if mine_id else ""
        section_name = db.get_section_name(section_id) if section_id else ""
        summary_rows.append((pid, name, base_budget, st["have"], st["need"], st["diff"], st["stage"], bool(out_of_budget), mine_name, section_name))

    # сначала создадим все листы проектов, чтобы в сводной можно было проставить корректные гиперссылки
    sheet_map: dict[int, str] = {}
    for pid, name, *_ in summary_rows:
        _create_project_sheet(wb, pid, name or f"Проект {pid}", used_names, sheet_map, head_fill, head_font, right)

    # Заполнение сводной + гиперссылки
    for row_idx, (pid, name, base, have, need, diff, stage, out_of_budget, mine_name, section_name) in enumerate(summary_rows, start=2):
        link_cell = _write_summary_cell(ws_sum, row_idx, 1, name or f"Проект {pid}")
        target_sheet = sheet_map.get(pid)
        if target_sheet:
            link_cell.hyperlink = f"#'{target_sheet}'!A1"
            link_cell.style = "Hyperlink"
        _write_summary_cell(ws_sum, row_idx, 2, base, money=True, right=right)
        _write_summary_cell(ws_sum, row_idx, 3, have, money=True, right=right)
        _write_summary_cell(ws_sum, row_idx, 4, need, money=True, right=right)
        _write_summary_cell(ws_sum, row_idx, 5, diff, money=True, right=right)
        ws_sum.cell(row=row_idx, column=6, value=stage)
        ws_sum.cell(row=row_idx, column=7, value="Вне бюджета" if out_of_budget else "Бюджет")
        ws_sum.cell(row=row_idx, column=8, value=mine_name or "")
        ws_sum.cell(row=row_idx, column=9, value=section_name or "")
        for c in (8, 9):
            ws_sum.cell(row=row_idx, column=c).alignment = right

    _autosize(ws_sum)

    # Сохранение
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
    wb.save(xlsx_path)
    return os.path.abspath(xlsx_path)


if __name__ == "__main__":
    # Простой ручной тест: export_excel.py -> data/export_YYYYMMDD.xlsx
    ts = _dt.date.today().strftime("%Y%m%d")
    out = os.path.join("data", f"export_{ts}.xlsx")
    print("Export to:", export_to_excel(out))
